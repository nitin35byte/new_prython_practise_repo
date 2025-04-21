import statistics

import openpyxl

workbook= openpyxl.load_workbook("testdata.xlsx")
sheet= workbook.active
print(sheet)

row_count=sheet.max_row
column_count=sheet.max_column
print(row_count)
print(column_count)


for  row in range(1,row_count+1):
    for col in range(1,column_count+1):
        cell_value=sheet.cell(row=row , column=col).value
        print(cell_value)


import pandas as pd

df =pd.read_excel("testdata.xlsx")

n=df['name'].tolist()
print(n)



class readdata:

    # def __init__(self,driver):
    #     self.driver = driver
    """
    This is a instance method
    """
    def readt(self,filename):
        workbook= openpyxl.load_workbook(filename)
        sheet=workbook.active
        row_count=sheet.max_row
        column_count=sheet.max_column
        print(row_count)
        for row in range(1, row_count + 1):
            for col in range(1, column_count + 1):
                cell_value = sheet.cell(row=row, column=col).value
                print(cell_value)


handler= readdata()

print(handler.readt("testdata.xlsx"))


class Ereaddata:
    """
    This is a static method
    """
    def readt(filename):
        workbook= openpyxl.load_workbook(filename)
        sheet=workbook.active
        row_count=sheet.max_row
        column_count=sheet.max_column
        print(row_count)
        for row in range(1, row_count + 1):
            for col in range(1, column_count + 1):
                cell_value = sheet.cell(row=row, column=col).value
                print(cell_value)


print(Ereaddata.readt("testdata.xlsx"))