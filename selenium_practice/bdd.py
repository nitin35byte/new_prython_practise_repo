import openpyxl

workbook=openpyxl.load_workbook("test_data.xlsx")
sheet=workbook.active
maxrow=sheet.max_row
max_col=sheet.max_column
print(maxrow)
print(max_col)

for row in  range(2 , maxrow+1):
    for col in range(2, max_col+1):
        s=sheet.cell(row=row , column=col).value
        print(s)

        # read all the data and print
        data= sheet.cell(row=row , column=col).value
        print(data)

        #Read single values and print
        # cell_value=sheet.cell(row=row , column=1).value
        # cell_value1 = sheet.cell(row=row, column=2).value
        # print(cell_value1 , cell_value)


with open("random_test.txt" , 'w') as f:
    f.write("Nitin\n"
            'Nitish\n'
            'Nishant\n')
    f.close()


with open("random_test.txt" , 'r')as f:
    data=[i.strip() for i in f.readlines()]

print(data[2])
