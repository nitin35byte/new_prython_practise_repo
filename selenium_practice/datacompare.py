import  openpyxl

workbook=openpyxl.load_workbook('testdata.xlsx')
sheet=workbook.active
print(sheet)

row_count=sheet.max_row
col_count=sheet.max_column


print(f"max row is {row_count} and max column is {col_count}")

for row in range(2 , row_count+1):
    cell=sheet.cell(row=row, column=1).value
    print(cell)

if cell == sorted(cell):
    print("data is sorted ")
else:
    print("data is not sorted")


import pandas as pd
df = pd.read_excel("testdata.xlsx" , index_col=False)
print(df)
names=df['name'].tolist()

# for name in names:
if names == sorted(names):
    print("name is sorted ")
else:
    print("data is not sorted")

# Check sorted order and identify unsorted positions
unsorted_positions = [
    i + 1 for i in range(len(names) - 1) if names[i] > names[i + 1]
]

if not unsorted_positions:
    print("The names are sorted.")
else:
    print(f"The names are not sorted. Unsorted positions: {unsorted_positions}")
